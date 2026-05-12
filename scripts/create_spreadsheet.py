"""
Create the Pudim Caramelo v2 spreadsheet with normalized data model.

Migrates data from the original flat spreadsheet into a structured format
with separate tabs for products, suppliers, purchases, etc.

Usage:
    python3 scripts/create_spreadsheet.py

Output:
    output/Pudim_Caramelo_v2.xlsx
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# --- Style definitions ---

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4A4A8A", end_color="4A4A8A", fill_type="solid")
ALI_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # green tint
FOR_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # orange tint
EMB_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # blue tint
EQP_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # purple tint
CALC_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # yellow tint
SECTION_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BRL_FORMAT = 'R$ #,##0.00'
PCT_FORMAT = '0.0%'


def style_header(ws, num_cols):
    """Apply header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, num_rows, num_cols, category_fill=None):
    """Apply borders and optional category fill to data rows."""
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if category_fill:
                cell.fill = category_fill


def auto_width(ws, num_cols, min_width=12, max_width=40):
    """Auto-adjust column widths based on content."""
    for col in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def apply_currency_format(ws, col, start_row=2, end_row=None):
    """Apply BRL currency format to a column."""
    if end_row is None:
        end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            cell.number_format = BRL_FORMAT


# =============================================================================
# DATA
# =============================================================================

PRODUTOS = [
    # ALI - Alimentos
    ("ALI-001", "LEITE CONDENSADO 395G", "UN", ""),
    ("ALI-002", "AÇÚCAR REFINADO 1KG", "KG", ""),
    ("ALI-003", "LEITE INTEGRAL 1L", "L", ""),
    ("ALI-004", "OVO (UNIDADE)", "UN", "Comprado em pente de 30"),
    # FOR - Formas
    ("FOR-001", "FORMA PUDIM GRANDE 500ML", "UN", "Plastilânia"),
    ("FOR-002", "FORMA PUDIM FAMÍLIA 1,1L", "UN", ""),
    ("FOR-003", "FORMA PUDIM MÉDIA 250ML", "UN", "Plastilânia"),
    ("FOR-004", "FORMA PUDIM PEQUENA 150ML", "UN", "Plastilânia"),
    ("FOR-005", "FORMA QUADRADA 220ML", "UN", "Plastilânia"),
    # EMB - Embalagens
    ("EMB-001", "SACOLA KRAFT GG (35X30X18)", "UN", ""),
    ("EMB-002", "SACOLA KRAFT G (31X26X13)", "UN", ""),
    ("EMB-003", "SACO CELOFANE GG (45X59)", "UN", ""),
    ("EMB-004", "SACO CELOFANE G (30X44)", "UN", ""),
    ("EMB-005", "FITA PRODUTO ARTESANAL", "M", ""),
    ("EMB-006", "COLHER ROXA", "UN", ""),
    ("EMB-007", "SAQUINHO P/ COLHER", "UN", ""),
    ("EMB-008", "ETIQUETA REDONDA TAMPA VINIL (7,5x7,5)", "FOLHA", ""),
    ("EMB-009", "ADESIVO CACHORRO CARAMELO", "FOLHA", ""),
    ("EMB-010", "ADESIVO BANDEIRA CARAMELO", "FOLHA", ""),
    ("EMB-011", "ADESIVO ME VÊ 2 FATIAS", "FOLHA", ""),
    ("EMB-012", "SACOLA KRAFT (24X15X25)", "UN", ""),
    ("EMB-013", "SACO PLÁSTICO POLIPROPILENO (30X44)", "UN", ""),
    ("EMB-014", "FITA CETIM 15MM ROXO", "M", ""),
    ("EMB-015", "FITA CETIM 15MM LILÁS", "M", ""),
    ("EMB-016", "BARBANTE SISAL", "M", ""),
    ("EMB-017", "BARBANTE VERMELHO E BRANCO", "M", ""),
    ("EMB-018", "ETIQUETA REDONDA TAMPA PAPEL (7,5x7,5)", "FOLHA", ""),
    ("EMB-019", "SACOLA KRAFT (31X17X30)", "UN", ""),
    ("EMB-020", "SACO PLÁSTICO POLIPROPILENO (35X50)", "UN", ""),
    # EQP - Equipamentos
    ("EQP-001", "GRADE METAL 46X26X3", "UN", ""),
    ("EQP-002", "LUVA SILICONE", "UN", ""),
    ("EQP-003", "PANELINHA MAX TRANSP C/TAMPA", "UN", ""),
    ("EQP-004", "GRADE DE RESFRIAMENTO", "UN", ""),
]

FORNECEDORES = [
    ("FORN-001", "Apoio Mineiro", "Supermercado", "", ""),
    ("FORN-002", "Supernosso", "Supermercado", "", ""),
    ("FORN-003", "Maria Chocolate", "Loja", "", ""),
    ("FORN-004", "1001 Festas", "Loja", "", ""),
    ("FORN-005", "Impacto", "Loja", "", ""),
    ("FORN-006", "Feira", "Feira", "", ""),
    ("FORN-007", "iFood", "Online", "", ""),
    ("FORN-008", "QualiGraf", "Gráfica", "", ""),
    ("FORN-009", "Distribuidora Belopack", "Loja", "", ""),
    ("FORN-010", "Mercado Livre (EmbalaiadoSP)", "Online", "", ""),
    ("FORN-011", "Gato Preto (Mercado Central)", "Loja", "", ""),
]

# Compras: migrated from current insumos tabs
# (ID, Data, Produto_ID, Fornecedor_ID, Marca, Qtde_Emb, Unid_por_Emb, Preco_Total, Notas)
# Total_Unidades and Preco_Unitario are calculated columns
COMPRAS = [
    # Alimentos
    ("C-001", date(2025, 1, 1), "ALI-002", "FORN-003", "UNIÃO", 1, 1, 6.50, "Migrado da planilha original"),
    ("C-002", date(2025, 1, 1), "ALI-001", "FORN-002", "CEMIL", 1, 1, 6.29, "Migrado da planilha original"),
    ("C-003", date(2025, 1, 1), "ALI-003", "FORN-002", "ITAMBÉ", 1, 1, 4.79, "Migrado da planilha original"),
    ("C-004", date(2025, 1, 1), "ALI-004", "FORN-002", "ASA", 1, 30, 22.99, "Ovo branco"),
    ("C-005", date(2025, 1, 1), "ALI-001", "FORN-004", "CEMIL", 1, 1, 5.99, "Migrado da planilha original"),
    ("C-006", date(2025, 1, 1), "ALI-004", "FORN-006", "FEIRA", 1, 30, 35.00, "Ovo vermelho"),
    ("C-007", date(2025, 1, 1), "ALI-001", "FORN-001", "CEMIL", 1, 1, 5.49, "Migrado da planilha original"),
    ("C-008", date(2025, 1, 1), "ALI-003", "FORN-001", "PORTO ALEGRE", 1, 1, 3.59, "Migrado da planilha original"),
    ("C-009", date(2025, 1, 1), "ALI-004", "FORN-006", "FEIRA", 1, 30, 30.00, "Ovo vermelho - Promoção"),
    ("C-010", date(2025, 1, 1), "ALI-002", "FORN-007", "UNIÃO", 1, 1, 5.91, "Migrado da planilha original"),
    # Formas
    ("C-011", date(2025, 1, 1), "FOR-001", "FORN-003", "PLASTILÂNIA", 1, 6, 21.90, "Migrado da planilha original"),
    ("C-012", date(2025, 1, 1), "FOR-002", "FORN-005", "PLASTILÂNIA", 1, 5, 34.80, "Migrado da planilha original"),
    ("C-013", date(2025, 1, 1), "FOR-003", "FORN-005", "PLASTILÂNIA", 1, 8, 21.90, "Migrado da planilha original"),
    ("C-014", date(2025, 1, 1), "FOR-004", "FORN-005", "PLASTILÂNIA", 1, 10, 13.50, "Migrado da planilha original"),
    ("C-015", date(2025, 1, 1), "FOR-002", "FORN-003", "BLUE STAR", 1, 5, 37.00, "Migrado da planilha original"),
    ("C-016", date(2025, 1, 1), "FOR-005", "FORN-003", "PLASTILÂNIA", 1, 10, 14.90, "Migrado da planilha original"),
    ("C-017", date(2025, 1, 1), "FOR-001", "FORN-009", "PLASTILÂNIA", 1, 6, 23.90, "Migrado da planilha original"),
    ("C-018", date(2025, 1, 1), "FOR-001", "FORN-010", "PLASTILÂNIA", 1, 120, 395.00, "Compra em volume"),
    ("C-019", date(2025, 1, 1), "FOR-002", "FORN-010", "PLASTILÂNIA", 1, 50, 316.29, "Compra em volume"),
    # Embalagens
    ("C-020", date(2025, 1, 1), "EMB-001", "FORN-005", "VINCO", 1, 25, 54.75, "Migrado da planilha original"),
    ("C-021", date(2025, 1, 1), "EMB-002", "FORN-005", "VINCO", 1, 25, 49.80, "Migrado da planilha original"),
    ("C-022", date(2025, 1, 1), "EMB-003", "FORN-003", "", 1, 25, 27.00, "Migrado da planilha original"),
    ("C-023", date(2025, 1, 1), "EMB-004", "FORN-003", "CROMUS", 1, 25, 14.60, "Migrado da planilha original"),
    ("C-024", date(2025, 1, 1), "EMB-005", "FORN-003", "PROGRESSO", 1, 10, 9.50, "Migrado da planilha original"),
    ("C-025", date(2025, 1, 1), "EMB-006", "FORN-005", "CROPAC", 1, 50, 8.90, "Migrado da planilha original"),
    ("C-026", date(2025, 1, 1), "EMB-007", "FORN-005", "", 1, 100, 1.99, "Migrado da planilha original"),
    ("C-027", date(2025, 1, 1), "EMB-008", "FORN-008", "VINIL BRILHO", 1, 20, 20.00, "Migrado - qty was 18 folhas, updated to 20"),
    ("C-028", date(2025, 1, 1), "EMB-009", "FORN-008", "VINIL FOSCO", 1, 40, 22.00, "Migrado da planilha original"),
    ("C-029", date(2025, 1, 1), "EMB-010", "FORN-008", "VINIL FOSCO", 1, 62, 22.00, "Migrado da planilha original"),
    ("C-030", date(2025, 1, 1), "EMB-011", "FORN-008", "VINIL FOSCO", 1, 47, 22.00, "Migrado da planilha original"),
    ("C-031", date(2025, 1, 1), "EMB-012", "FORN-011", "", 1, 50, 72.00, "Migrado da planilha original"),
    ("C-032", date(2025, 1, 1), "EMB-013", "FORN-009", "", 1, 100, 19.50, "Migrado da planilha original"),
    ("C-033", date(2025, 1, 1), "EMB-014", "FORN-009", "PROGRESSO", 1, 10, 6.25, "Migrado da planilha original"),
    ("C-034", date(2025, 1, 1), "EMB-015", "FORN-009", "PROGRESSO", 1, 10, 6.25, "Migrado da planilha original"),
    ("C-035", date(2025, 1, 1), "EMB-016", "FORN-004", "TUUT", 1, 36.5, 11.99, "Migrado da planilha original"),
    ("C-036", date(2025, 1, 1), "EMB-017", "FORN-004", "TUUT", 1, 100, 13.99, "Migrado da planilha original"),
    ("C-037", date(2025, 1, 1), "EMB-018", "FORN-008", "PAPEL", 1, 20, 16.00, "Migrado da planilha original"),
    ("C-038", date(2025, 1, 1), "EMB-019", "FORN-011", "", 1, 50, 136.00, "Migrado da planilha original"),
    ("C-039", date(2025, 1, 1), "EMB-020", "FORN-009", "", 1, 100, 20.00, "Migrado da planilha original"),
    # Equipamentos
    ("C-040", date(2025, 1, 1), "EQP-001", "FORN-003", "", 1, 1, 69.00, "Migrado da planilha original"),
    ("C-041", date(2025, 1, 1), "EQP-002", "FORN-003", "", 1, 1, 48.00, "Migrado da planilha original"),
    ("C-042", date(2025, 1, 1), "EQP-003", "FORN-003", "", 1, 2, 23.90, "Migrado da planilha original"),
    ("C-043", date(2025, 1, 1), "EQP-004", "FORN-004", "", 1, 2, 24.90, "Migrado da planilha original"),
]

TAMANHOS = [
    ("TAM-001", "Médio 500g", 0.5, 500, 2, "Fornada", 40.00, ""),
    ("TAM-002", "Grande 1Kg", 1.0, 1100, 1, "Pronta Entrega", 90.00, ""),
    ("TAM-003", "Quadrado 200g", 0.22, 220, 5, "Pronta Entrega", None, "Preço a definir"),
]

RECEITA = [
    ("ALI-002", 0.1),   # Açúcar 100g
    ("ALI-001", 1),     # 1 caixa leite condensado
    ("ALI-003", 0.395), # 395ml leite
    ("ALI-004", 4),     # 4 ovos
]

# Embalagens por tamanho: (Tamanho_ID, Produto_ID, Qtde)
EMBALAGENS_POR_TAMANHO = [
    # TAM-001 — Médio 500g
    ("TAM-001", "FOR-001", 1),
    ("TAM-001", "EMB-012", 1),
    ("TAM-001", "EMB-013", 1),
    ("TAM-001", "EMB-017", 0.45),
    ("TAM-001", "EMB-018", 1),
    ("TAM-001", "EMB-009", 1),
    ("TAM-001", "EMB-010", 1),
    ("TAM-001", "EMB-011", 1),
    # TAM-002 — Grande 1Kg
    ("TAM-002", "FOR-002", 1),
    ("TAM-002", "EMB-019", 1),
    ("TAM-002", "EMB-003", 1),
    ("TAM-002", "EMB-017", 0.45),
    ("TAM-002", "EMB-018", 1),
    ("TAM-002", "EMB-009", 1),
    ("TAM-002", "EMB-010", 1),
    # TAM-003 — Quadrado 200g
    ("TAM-003", "FOR-005", 1),
    ("TAM-003", "EMB-012", 1),
    ("TAM-003", "EMB-013", 1),
    ("TAM-003", "EMB-017", 0.45),
    ("TAM-003", "EMB-018", 1),
    ("TAM-003", "EMB-009", 1),
    ("TAM-003", "EMB-010", 1),
    ("TAM-003", "EMB-011", 1),
]

FORNADAS = [
    ("FN-001", date(2025, 11, 24), date(2025, 11, 28), "TAM-001", 13, 13, 0, 40.00, "Fornada 01"),
    ("FN-002", date(2025, 12, 1), date(2025, 12, 5), "TAM-001", 11, 11, 0, 40.00, "Fornada 02"),
    ("FN-003", date(2025, 12, 1), date(2025, 12, 5), "TAM-001", 11, 11, 0, 40.00, "Fornada 03"),
    ("FN-004", date(2025, 12, 1), date(2025, 12, 5), "TAM-001", 11, 11, 0, 40.00, "Fornada 04"),
    ("FN-005", date(2025, 12, 1), date(2025, 12, 5), "TAM-001", 11, 11, 0, 40.00, "Fornada Especial de Natal"),
]

FLUXO_CAIXA = [
    ("FC-001", date(2025, 11, 24), "Entrada", 520.00, "Venda", "FN-001", "Venda 13 pudins fornada 01"),
    ("FC-002", date(2025, 11, 25), "Saída", 35.00, "Compra Insumos", "", "Ovos feira"),
    ("FC-003", date(2025, 12, 1), "Saída", 151.50, "Compra Embalagem", "", "Embalagens"),
    ("FC-004", date(2025, 12, 1), "Saída", 106.73, "Compra Embalagem", "", "Embalagens"),
    ("FC-005", date(2025, 12, 1), "Saída", 99.00, "Compra Embalagem", "", "Embalagens"),
    ("FC-006", date(2025, 12, 1), "Saída", 36.00, "Transporte", "", "Transporte"),
]


# =============================================================================
# BUILD WORKBOOK
# =============================================================================

def build_produtos(wb):
    ws = wb.create_sheet("Produtos")
    headers = ["ID", "Nome", "Unidade", "Notas"]
    ws.append(headers)
    style_header(ws, len(headers))

    for row_data in PRODUTOS:
        ws.append(row_data)

    # Apply category fills
    for row in range(2, ws.max_row + 1):
        prod_id = ws.cell(row=row, column=1).value or ""
        if prod_id.startswith("ALI-"):
            fill = ALI_FILL
        elif prod_id.startswith("FOR-"):
            fill = FOR_FILL
        elif prod_id.startswith("EMB-"):
            fill = EMB_FILL
        elif prod_id.startswith("EQP-"):
            fill = EQP_FILL
        else:
            fill = None
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = THIN_BORDER

    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.freeze_panes = "A2"


def build_fornecedores(wb):
    ws = wb.create_sheet("Fornecedores")
    headers = ["ID", "Nome", "Tipo", "Localização", "Notas"]
    ws.append(headers)
    style_header(ws, len(headers))

    for row_data in FORNECEDORES:
        ws.append(row_data)

    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    ws.freeze_panes = "A2"


def build_compras(wb):
    ws = wb.create_sheet("Compras")
    headers = [
        "ID", "Data", "Produto_ID", "Fornecedor_ID", "Marca",
        "Qtde_Embalagens", "Unidades_por_Embalagem", "Total_Unidades",
        "Preco_Total", "Preco_Unitario", "Notas"
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, c in enumerate(COMPRAS):
        row_num = i + 2
        # Write non-calculated fields
        ws.cell(row=row_num, column=1, value=c[0])   # ID
        ws.cell(row=row_num, column=2, value=c[1])   # Data
        ws.cell(row=row_num, column=2).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_num, column=3, value=c[2])   # Produto_ID
        ws.cell(row=row_num, column=4, value=c[3])   # Fornecedor_ID
        ws.cell(row=row_num, column=5, value=c[4])   # Marca
        ws.cell(row=row_num, column=6, value=c[5])   # Qtde_Embalagens
        ws.cell(row=row_num, column=7, value=c[6])   # Unidades_por_Embalagem
        # Total_Unidades = Qtde_Emb * Unid_por_Emb
        ws.cell(row=row_num, column=8, value=f"=F{row_num}*G{row_num}")
        ws.cell(row=row_num, column=9, value=c[7])   # Preco_Total
        ws.cell(row=row_num, column=9).number_format = BRL_FORMAT
        # Preco_Unitario = Preco_Total / Total_Unidades
        ws.cell(row=row_num, column=10, value=f"=IF(H{row_num}>0,I{row_num}/H{row_num},0)")
        ws.cell(row=row_num, column=10).number_format = BRL_FORMAT
        ws.cell(row=row_num, column=11, value=c[8])  # Notas

    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    ws.freeze_panes = "A2"


def build_tamanhos(wb):
    ws = wb.create_sheet("Tamanhos")
    headers = ["ID", "Nome", "Peso_KG", "Volume_ML", "Rendimento", "Canal", "Preco_Venda", "Notas"]
    ws.append(headers)
    style_header(ws, len(headers))

    for row_data in TAMANHOS:
        ws.append(row_data)

    apply_currency_format(ws, 7)
    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def build_receita(wb):
    ws = wb.create_sheet("Receita")
    headers = ["Produto_ID", "Nome_Produto", "Qtde", "Unidade"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, (prod_id, qtde) in enumerate(RECEITA):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=prod_id)
        # Nome_Produto = VLOOKUP from Produtos
        ws.cell(row=row_num, column=2, value=f'=VLOOKUP(A{row_num},Produtos!A:B,2,FALSE)')
        ws.cell(row=row_num, column=3, value=qtde)
        # Unidade = VLOOKUP from Produtos
        ws.cell(row=row_num, column=4, value=f'=VLOOKUP(A{row_num},Produtos!A:C,3,FALSE)')

    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def build_embalagens_por_tamanho(wb):
    ws = wb.create_sheet("Embalagens_Por_Tamanho")
    headers = ["Tamanho_ID", "Produto_ID", "Nome_Produto", "Qtde_Por_Unidade"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, (tam_id, prod_id, qtde) in enumerate(EMBALAGENS_POR_TAMANHO):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=tam_id)
        ws.cell(row=row_num, column=2, value=prod_id)
        ws.cell(row=row_num, column=3, value=f'=VLOOKUP(B{row_num},Produtos!A:B,2,FALSE)')
        ws.cell(row=row_num, column=4, value=qtde)

    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.freeze_panes = "A2"


def build_fornadas(wb):
    ws = wb.create_sheet("Fornadas")
    headers = [
        "ID", "Data_Inicio", "Data_Fim", "Tamanho_ID",
        "Qtde_Produzida", "Qtde_Vendida", "Qtde_Cortesia",
        "Preco_Venda_Unit", "Receita_Total", "Custo_Unit",
        "Custo_Total", "Lucro", "Notas"
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, fn in enumerate(FORNADAS):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=fn[0])   # ID
        ws.cell(row=row_num, column=2, value=fn[1])   # Data_Inicio
        ws.cell(row=row_num, column=2).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_num, column=3, value=fn[2])   # Data_Fim
        ws.cell(row=row_num, column=3).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_num, column=4, value=fn[3])   # Tamanho_ID
        ws.cell(row=row_num, column=5, value=fn[4])   # Qtde_Produzida
        ws.cell(row=row_num, column=6, value=fn[5])   # Qtde_Vendida
        ws.cell(row=row_num, column=7, value=fn[6])   # Qtde_Cortesia
        ws.cell(row=row_num, column=8, value=fn[7])   # Preco_Venda_Unit
        ws.cell(row=row_num, column=8).number_format = BRL_FORMAT
        # Receita_Total = Vendida * Preço
        ws.cell(row=row_num, column=9, value=f"=F{row_num}*H{row_num}")
        ws.cell(row=row_num, column=9).number_format = BRL_FORMAT
        # Custo_Unit — placeholder, will be filled by Apps Script with Ficha Técnica reference
        ws.cell(row=row_num, column=10, value="")
        ws.cell(row=row_num, column=10).number_format = BRL_FORMAT
        # Custo_Total = Custo_Unit * Qtde_Produzida
        ws.cell(row=row_num, column=11, value=f"=IF(J{row_num}<>\"\",J{row_num}*E{row_num},\"\")")
        ws.cell(row=row_num, column=11).number_format = BRL_FORMAT
        # Lucro = Receita - Custo
        ws.cell(row=row_num, column=12, value=f"=IF(K{row_num}<>\"\",I{row_num}-K{row_num},\"\")")
        ws.cell(row=row_num, column=12).number_format = BRL_FORMAT
        ws.cell(row=row_num, column=13, value=fn[8])  # Notas

    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def build_fluxo_caixa(wb):
    ws = wb.create_sheet("Fluxo_Caixa")
    headers = ["ID", "Data", "Tipo", "Valor", "Categoria", "Referencia", "Descricao"]
    ws.append(headers)
    style_header(ws, len(headers))

    for row_data in FLUXO_CAIXA:
        ws.append(row_data)

    # Format dates and currency
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = 'DD/MM/YYYY'
        ws.cell(row=row, column=4).number_format = BRL_FORMAT

    style_data_rows(ws, ws.max_row, len(headers))
    auto_width(ws, len(headers))
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.freeze_panes = "A2"


def build_ficha_tecnica_placeholder(wb):
    """Create Ficha_Tecnica tab with structure. Formulas added by Apps Script."""
    ws = wb.create_sheet("Ficha_Tecnica")

    # Title
    ws.cell(row=1, column=1, value="FICHA TÉCNICA — Custos por Tamanho")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws.cell(row=3, column=1, value="As fórmulas desta aba serão adicionadas pelo Apps Script.")
    ws.cell(row=3, column=1).font = Font(italic=True, color="999999")
    ws.cell(row=4, column=1, value="Execute o script após importar esta planilha no Google Sheets.")
    ws.cell(row=4, column=1).font = Font(italic=True, color="999999")


def build_calculadora_placeholder(wb):
    """Create Calculadora tab with structure."""
    ws = wb.create_sheet("Calculadora")
    ws.cell(row=1, column=1, value="CALCULADORA DE PREÇO E MARGEM")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="As fórmulas desta aba serão adicionadas pelo Apps Script.")
    ws.cell(row=3, column=1).font = Font(italic=True, color="999999")


def build_producao_placeholder(wb):
    """Create Produção tab with structure."""
    ws = wb.create_sheet("Produção")
    ws.cell(row=1, column=1, value="PLANEJAMENTO DE PRODUÇÃO")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="As fórmulas desta aba serão adicionadas pelo Apps Script.")
    ws.cell(row=3, column=1).font = Font(italic=True, color="999999")


def build_comparativo_placeholder(wb):
    """Create Comparativo_Fornecedores tab."""
    ws = wb.create_sheet("Comparativo_Fornecedores")
    ws.cell(row=1, column=1, value="COMPARATIVO DE FORNECEDORES")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="As fórmulas desta aba serão adicionadas pelo Apps Script.")
    ws.cell(row=3, column=1).font = Font(italic=True, color="999999")


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Data tabs
    build_produtos(wb)
    build_fornecedores(wb)
    build_compras(wb)
    build_tamanhos(wb)
    build_receita(wb)
    build_embalagens_por_tamanho(wb)
    build_fornadas(wb)
    build_fluxo_caixa(wb)

    # Calculation tabs (placeholders — formulas via Apps Script)
    build_ficha_tecnica_placeholder(wb)
    build_calculadora_placeholder(wb)
    build_producao_placeholder(wb)
    build_comparativo_placeholder(wb)

    # Save
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "Pudim_Caramelo_v2.xlsx")
    wb.save(output_path)
    print(f"Spreadsheet created: {output_path}")
    print(f"  - {len(PRODUTOS)} products")
    print(f"  - {len(FORNECEDORES)} suppliers")
    print(f"  - {len(COMPRAS)} purchase records")
    print(f"  - {len(TAMANHOS)} sizes")
    print(f"  - {len(RECEITA)} recipe ingredients")
    print(f"  - {len(EMBALAGENS_POR_TAMANHO)} packaging entries")
    print(f"  - {len(FORNADAS)} batch records")
    print(f"  - {len(FLUXO_CAIXA)} cash flow entries")


if __name__ == "__main__":
    main()
